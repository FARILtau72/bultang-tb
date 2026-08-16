"""
Attendance (Absensi) management business logic.
"""

from io import BytesIO

import pandas as pd
from sqlalchemy import text
from sqlalchemy.exc import IntegrityError

from app.core.config import JURUSAN, STATUS_OPTIONS
from app.core.db import get_engine
from app.core.utils import get_now_time_str, get_today_str, get_now_wib


def process_scan_qr(kode_qr: str, status: str = "HADIR") -> dict:
    """
    Process QR code scan for attendance.
    
    Args:
        kode_qr: QR code identifier
        status: Attendance status (default: HADIR)
        
    Returns:
        Dictionary with operation result
    """
    now_wib = get_now_wib()
    if now_wib.weekday() != 0:  # 0 = Monday
        return {
            "status": "error",
            "message": "Absensi hanya diperbolehkan pada hari Senin.",
        }

    if not kode_qr:
        return {"status": "error", "message": "Kode QR tidak terbaca."}

    status_clean = (status or "HADIR").upper()
    if status_clean not in STATUS_OPTIONS:
        status_clean = "HADIR"

    today = get_today_str()
    now_time = get_now_time_str()
    engine = get_engine()

    try:
        with engine.begin() as conn:
            siswa = conn.execute(
                text(
                    """
                    SELECT id, nama, kelas, jurusan
                    FROM siswa
                    WHERE kode_qr = :kode_qr
                    """
                ),
                {"kode_qr": kode_qr},
            ).mappings().first()

            if not siswa:
                return {"status": "error", "message": "QR Code siswa tidak ditemukan."}

            existing = conn.execute(
                text(
                    """
                    SELECT waktu, status FROM absensi
                    WHERE siswa_id = :siswa_id AND tanggal = :tanggal
                    """
                ),
                {"siswa_id": siswa["id"], "tanggal": today},
            ).mappings().first()
            
            if existing:
                return {
                    "status": "already",
                    "nama": siswa["nama"],
                    "waktu": existing["waktu"],
                    "message": f"{siswa['nama']} sudah tercatat hari ini ({existing['waktu']}).",
                }

            conn.execute(
                text(
                    """
                    INSERT INTO absensi (siswa_id, nama, kelas, jurusan, waktu, tanggal, status)
                    VALUES (:siswa_id, :nama, :kelas, :jurusan, :waktu, :tanggal, :status)
                    """
                ),
                {
                    "siswa_id": siswa["id"],
                    "nama": siswa["nama"],
                    "kelas": siswa["kelas"],
                    "jurusan": siswa["jurusan"],
                    "waktu": now_time,
                    "tanggal": today,
                    "status": status_clean,
                },
            )
    except IntegrityError:
        with engine.connect() as conn:
            existing = conn.execute(
                text(
                    "SELECT waktu FROM absensi WHERE siswa_id = :siswa_id AND tanggal = :tanggal"
                ),
                {"siswa_id": siswa["id"], "tanggal": today},
            ).mappings().first()
        return {
            "status": "already",
            "nama": siswa["nama"],
            "waktu": existing["waktu"] if existing else "-",
            "message": f"{siswa['nama']} sudah tercatat hari ini.",
        }

    return {
        "status": "success",
        "nama": siswa["nama"],
        "kelas": siswa["kelas"],
        "jurusan": siswa["jurusan"],
        "waktu": now_time,
        "message": f"Absensi {siswa['nama']} berhasil.",
    }


def get_dashboard_metrics() -> dict:
    """
    Get dashboard metrics for today's attendance.
    
    Returns:
        Dictionary with attendance metrics
    """
    today = get_today_str()
    engine = get_engine()
    
    with engine.connect() as conn:
        total_siswa = int(
            conn.execute(text("SELECT COUNT(*) AS count FROM siswa")).mappings().first()["count"]
        )
        hadir_hari_ini = int(
            conn.execute(
                text(
                    """
                    SELECT COUNT(*) AS count
                    FROM absensi
                    WHERE tanggal = :tanggal AND status = 'HADIR'
                    """
                ),
                {"tanggal": today},
            ).mappings().first()["count"]
        )

    belum_hadir = max(total_siswa - hadir_hari_ini, 0)
    persentase = round((hadir_hari_ini / total_siswa) * 100, 2) if total_siswa else 0

    return {
        "tanggal": today,
        "total_siswa": total_siswa,
        "hadir_hari_ini": hadir_hari_ini,
        "belum_hadir": belum_hadir,
        "persentase": persentase,
    }


def get_today_jurusan_summary() -> list[dict]:
    """
    Get attendance summary by major for today.
    
    Returns:
        List of attendance summaries per major
    """
    today = get_today_str()
    engine = get_engine()
    
    with engine.connect() as conn:
        rows = conn.execute(
            text(
                """
                SELECT
                    s.jurusan,
                    COUNT(s.id) AS total_siswa,
                    SUM(CASE WHEN a.id IS NOT NULL THEN 1 ELSE 0 END) AS hadir_count
                FROM siswa s
                LEFT JOIN absensi a
                    ON a.siswa_id = s.id
                   AND a.tanggal = :tanggal
                   AND a.status = 'HADIR'
                GROUP BY s.jurusan
                ORDER BY s.jurusan
                """
            ),
            {"tanggal": today},
        ).mappings().all()

    result: list[dict] = []
    for jurusan in JURUSAN:
        row = next((item for item in rows if item["jurusan"] == jurusan), None)
        result.append(
            {
                "jurusan": jurusan,
                "total_siswa": row["total_siswa"] if row else 0,
                "hadir_count": row["hadir_count"] if row else 0,
            }
        )
    return result


def get_recent_absensi(limit: int = 10) -> list[dict]:
    """
    Get recent attendance records.
    
    Args:
        limit: Number of records to retrieve
        
    Returns:
        List of recent attendance records
    """
    limit_value = max(1, int(limit))
    engine = get_engine()
    
    with engine.connect() as conn:
        rows = conn.execute(
            text(
                f"""
                SELECT id, nama, kelas, jurusan, waktu, tanggal, status
                FROM absensi
                ORDER BY tanggal DESC, waktu DESC, id DESC
                LIMIT {limit_value}
                """
            )
        ).mappings().all()

    return [dict(row) for row in rows]


def get_rekap_absensi(
    start_date: str | None = None,
    end_date: str | None = None,
    jurusan: str | None = None,
    kelas: str | None = None,
    status: str | None = None,
) -> list[dict]:
    """
    Get attendance records with filtering.
    
    Args:
        start_date: Start date (YYYY-MM-DD)
        end_date: End date (YYYY-MM-DD)
        jurusan: Filter by major
        kelas: Filter by class
        status: Filter by status
        
    Returns:
        List of attendance records
    """
    query = """
        SELECT id, nama, kelas, jurusan, waktu, tanggal, status
        FROM absensi
        WHERE 1=1
    """
    params: dict = {}

    if start_date:
        query += " AND tanggal >= :start_date"
        params["start_date"] = start_date
    if end_date:
        query += " AND tanggal <= :end_date"
        params["end_date"] = end_date
    if jurusan:
        query += " AND jurusan = :jurusan"
        params["jurusan"] = jurusan.upper()
    if kelas:
        query += " AND kelas = :kelas"
        params["kelas"] = kelas.upper()
    if status:
        query += " AND status = :status"
        params["status"] = status.upper()

    query += " ORDER BY tanggal DESC, waktu DESC, id DESC"

    engine = get_engine()
    with engine.connect() as conn:
        rows = conn.execute(text(query), params).mappings().all()

    return [dict(row) for row in rows]


def get_rekap_stats_per_kelas(
    start_date: str | None = None,
    end_date: str | None = None,
    jurusan: str | None = None,
    kelas: str | None = None,
) -> list[dict]:
    """
    Get attendance statistics per class.
    
    Args:
        start_date: Start date (YYYY-MM-DD)
        end_date: End date (YYYY-MM-DD)
        jurusan: Filter by major
        kelas: Filter by class
        
    Returns:
        List of statistics per class
    """
    query = """
        SELECT
            kelas,
            COUNT(*) AS total_absensi,
            SUM(CASE WHEN status = 'HADIR' THEN 1 ELSE 0 END) AS hadir,
            SUM(CASE WHEN status = 'IZIN' THEN 1 ELSE 0 END) AS izin,
            SUM(CASE WHEN status = 'SAKIT' THEN 1 ELSE 0 END) AS sakit,
            SUM(CASE WHEN status = 'ALPHA' THEN 1 ELSE 0 END) AS alpha
        FROM absensi
        WHERE 1=1
    """
    params: dict = {}

    if start_date:
        query += " AND tanggal >= :start_date"
        params["start_date"] = start_date
    if end_date:
        query += " AND tanggal <= :end_date"
        params["end_date"] = end_date
    if jurusan:
        query += " AND jurusan = :jurusan"
        params["jurusan"] = jurusan.upper()
    if kelas:
        query += " AND kelas = :kelas"
        params["kelas"] = kelas.upper()

    query += " GROUP BY kelas ORDER BY kelas"

    engine = get_engine()
    with engine.connect() as conn:
        rows = conn.execute(text(query), params).mappings().all()

    return [dict(row) for row in rows]


def get_rekap_akumulasi_siswa(
    start_date: str | None = None,
    end_date: str | None = None,
    jurusan: str | None = None,
    kelas: str | None = None,
) -> list[dict]:
    """
    Get accumulated attendance statistics per student.
    
    Returns list of dicts with keys:
    no, nama, kelas, jurusan, hadir, izin, sakit, alpha, total_pertemuan
    """
    query = """
        SELECT
            s.id AS siswa_id,
            s.nama,
            s.kelas,
            s.jurusan,
            COALESCE(SUM(CASE WHEN a.status = 'HADIR' THEN 1 ELSE 0 END), 0) AS hadir,
            COALESCE(SUM(CASE WHEN a.status = 'IZIN' THEN 1 ELSE 0 END), 0) AS izin,
            COALESCE(SUM(CASE WHEN a.status = 'SAKIT' THEN 1 ELSE 0 END), 0) AS sakit,
            COALESCE(SUM(CASE WHEN a.status = 'ALPHA' THEN 1 ELSE 0 END), 0) AS alpha,
            COUNT(a.id) AS total_pertemuan
        FROM siswa s
        LEFT JOIN absensi a ON s.id = a.siswa_id
    """
    params: dict = {}
    join_conditions = []
    if start_date:
        join_conditions.append("a.tanggal >= :start_date")
        params["start_date"] = start_date
    if end_date:
        join_conditions.append("a.tanggal <= :end_date")
        params["end_date"] = end_date

    if join_conditions:
        query += " AND " + " AND ".join(join_conditions)

    where_conditions = []
    if jurusan:
        where_conditions.append("s.jurusan = :jurusan")
        params["jurusan"] = jurusan.upper()
    if kelas:
        where_conditions.append("s.kelas = :kelas")
        params["kelas"] = kelas.upper()

    if where_conditions:
        query += " WHERE " + " AND ".join(where_conditions)

    query += " GROUP BY s.id, s.nama, s.kelas, s.jurusan ORDER BY s.kelas, s.nama"

    engine = get_engine()
    with engine.connect() as conn:
        rows = conn.execute(text(query), params).mappings().all()

    result = []
    for idx, row in enumerate(rows, start=1):
        d = dict(row)
        d["no"] = idx
        result.append(d)

    return result


def export_rekap_excel(
    start_date: str | None = None,
    end_date: str | None = None,
    jurusan: str | None = None,
    kelas: str | None = None,
    status: str | None = None,
) -> BytesIO:
    """
    Export accumulated student attendance records to formatted Excel.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from app.core.utils import format_tanggal_display

    rows = get_rekap_akumulasi_siswa(start_date, end_date, jurusan, kelas)

    if start_date and end_date:
        periode_str = f"{format_tanggal_display(start_date)} - {format_tanggal_display(end_date)}"
    elif start_date:
        periode_str = f"Mulai {format_tanggal_display(start_date)}"
    elif end_date:
        periode_str = f"Hingga {format_tanggal_display(end_date)}"
    else:
        periode_str = "Semua Periode"

    wb = Workbook()
    ws = wb.active
    ws.title = "Rekap Akumulasi"

    # Title & Subtitle
    ws.merge_cells("A1:H1")
    ws["A1"] = "REKAPITULASI AKUMULASI ABSENSI SISWA"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="1E293B")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("A2:H2")
    ws["A2"] = f"Periode: {periode_str} (Khusus Hari Senin)"
    ws["A2"].font = Font(name="Calibri", size=11, italic=True, color="475569")
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")

    # Table Header
    headers = [
        "No",
        "Nama Siswa",
        "Kelas",
        "Hadir",
        "Izin",
        "Sakit",
        "Alpha",
        "Total Pertemuan",
    ]
    ws.append([])  # Row 3 empty
    ws.append(headers)  # Row 4

    header_fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style="thin", color="CBD5E1"),
        right=Side(style="thin", color="CBD5E1"),
        top=Side(style="thin", color="CBD5E1"),
        bottom=Side(style="thin", color="CBD5E1"),
    )

    for col_num in range(1, 9):
        cell = ws.cell(row=4, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    # Data Rows
    tot_hadir = tot_izin = tot_sakit = tot_alpha = tot_pertemuan = 0
    row_idx = 5
    for r in rows:
        h = int(r["hadir"])
        i = int(r["izin"])
        s = int(r["sakit"])
        a = int(r["alpha"])
        tp = int(r["total_pertemuan"])

        tot_hadir += h
        tot_izin += i
        tot_sakit += s
        tot_alpha += a
        tot_pertemuan += tp

        row_data = [r["no"], r["nama"], r["kelas"], h, i, s, a, tp]
        ws.append(row_data)

        for col_num in range(1, 9):
            cell = ws.cell(row=row_idx, column=col_num)
            cell.border = thin_border
            if col_num in (1, 3, 4, 5, 6, 7, 8):
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")
        row_idx += 1

    # Total Row
    total_row = [
        "TOTAL AKUMULASI",
        "",
        "",
        tot_hadir,
        tot_izin,
        tot_sakit,
        tot_alpha,
        tot_pertemuan,
    ]
    ws.append(total_row)
    ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=3)

    total_fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
    total_font = Font(name="Calibri", size=11, bold=True, color="0F172A")

    for col_num in range(1, 9):
        cell = ws.cell(row=row_idx, column=col_num)
        cell.fill = total_fill
        cell.font = total_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Column Widths
    col_widths = {1: 6, 2: 32, 3: 14, 4: 10, 5: 10, 6: 10, 7: 10, 8: 18}
    for col_num, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col_num)].width = width

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def set_manual_status(siswa_id: int, status: str, tanggal: str | None = None) -> dict:
    """
    Set or update attendance status manually (by admin).
    
    If a record exists for the student on that date, update it.
    If not, insert a new record.
    
    Args:
        siswa_id: Student ID
        status: Attendance status (HADIR, IZIN, SAKIT, ALPHA)
        tanggal: Date string (defaults to today)
        
    Returns:
        Dict with operation result
    """
    from app.core.utils import get_today_str, get_now_time_str

    tanggal = tanggal or get_today_str()
    status_clean = (status or "").strip().upper()

    if status_clean not in STATUS_OPTIONS:
        return {"status": "error", "message": f"Status '{status}' tidak valid."}

    engine = get_engine()

    with engine.connect() as conn:
        siswa = conn.execute(
            text("SELECT id, nama, kelas, jurusan FROM siswa WHERE id = :siswa_id"),
            {"siswa_id": siswa_id},
        ).mappings().first()

    if not siswa:
        return {"status": "error", "message": "Siswa tidak ditemukan."}

    now_time = get_now_time_str()

    try:
        with engine.begin() as conn:
            existing = conn.execute(
                text(
                    "SELECT id FROM absensi WHERE siswa_id = :siswa_id AND tanggal = :tanggal"
                ),
                {"siswa_id": siswa_id, "tanggal": tanggal},
            ).mappings().first()

            if existing:
                conn.execute(
                    text(
                        """
                        UPDATE absensi 
                        SET status = :status, waktu = :waktu 
                        WHERE siswa_id = :siswa_id AND tanggal = :tanggal
                        """
                    ),
                    {
                        "status": status_clean,
                        "waktu": now_time,
                        "siswa_id": siswa_id,
                        "tanggal": tanggal,
                    },
                )
            else:
                conn.execute(
                    text(
                        """
                        INSERT INTO absensi (siswa_id, nama, kelas, jurusan, waktu, tanggal, status)
                        VALUES (:siswa_id, :nama, :kelas, :jurusan, :waktu, :tanggal, :status)
                        """
                    ),
                    {
                        "siswa_id": siswa_id,
                        "nama": siswa["nama"],
                        "kelas": siswa["kelas"],
                        "jurusan": siswa["jurusan"],
                        "waktu": now_time,
                        "tanggal": tanggal,
                        "status": status_clean,
                    },
                )
    except IntegrityError:
        return {"status": "error", "message": "Gagal menyimpan status absensi."}

    return {
        "status": "success",
        "message": f"Status {siswa['nama']} diubah menjadi {status_clean}.",
        "data": {
            "siswa_id": siswa_id,
            "nama": siswa["nama"],
            "status": status_clean,
            "tanggal": tanggal,
        },
    }


def get_today_attendance_map(tanggal: str | None = None) -> dict:
    """
    Get a mapping of siswa_id -> attendance info for a given date.
    
    Args:
        tanggal: Date string (defaults to today)
        
    Returns:
        Dict mapping siswa_id to {"status": str, "waktu": str}
    """
    from app.core.utils import get_today_str

    tanggal = tanggal or get_today_str()
    engine = get_engine()

    with engine.connect() as conn:
        rows = conn.execute(
            text(
                """
                SELECT siswa_id, status, waktu
                FROM absensi
                WHERE tanggal = :tanggal
                """
            ),
            {"tanggal": tanggal},
        ).mappings().all()

    return {
        row["siswa_id"]: {"status": row["status"], "waktu": row["waktu"]}
        for row in rows
    }

